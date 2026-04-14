#!/usr/bin/env python3
"""
apoHouze — Generiek EU Medicijnen Fetcher v2
=============================================
Gebruik: python3 fetch_eu_medicines.py <landcode>

Strategie per land:
  ALLE landen : EMA JSON (centraal vergunde EU-middelen, ~1546 records)
  ES           : AEMPS CIMA REST API (paginering, max 200 pagina's)
  DE           : BfArM Dringlichkeitsliste CSV + Pädiatrie-lijst
  AT           : AGES Lieferengpass + data.gv.at
  SE/NO/DK/FI  : Nationale tekortenlijsten (vergelijkbaar met AIFA carenti)
  CH           : Swissmedic directe download-links
  IE/PL/PT     : Nationale tekortenlijsten
"""

import sys, os, re, csv, time, subprocess, json, io, urllib.request, zipfile
from urllib.parse import urlencode

if len(sys.argv) < 2:
    print(f"Gebruik: python3 {sys.argv[0]} <landcode>"); sys.exit(1)

COUNTRY   = sys.argv[1].upper()
REPO_ROOT = os.getcwd()
TMP_DIR   = os.path.join(REPO_ROOT, "data", "_tmp")
OUTPUT    = os.path.join(TMP_DIR, f"{COUNTRY.lower()}_medicines.csv")
os.makedirs(TMP_DIR, exist_ok=True)
DEBUG = "--debug" in sys.argv

ATC_MAP = {
    "A02":"Stomach & Intestine","A03":"Stomach & Intestine","A04":"Stomach & Intestine",
    "A05":"Stomach & Intestine","A06":"Stomach & Intestine","A07":"Stomach & Intestine",
    "A08":"Stomach & Intestine","A09":"Stomach & Intestine","A10":"Diabetes",
    "A11":"Vitamins & Supplements","A12":"Vitamins & Supplements","A13":"Vitamins & Supplements",
    "A16":"Stomach & Intestine",
    "B01":"Anticoagulants","B02":"Heart & Blood Pressure","B03":"Vitamins & Supplements",
    "B05":"Heart & Blood Pressure","B06":"Heart & Blood Pressure",
    "C01":"Heart & Blood Pressure","C02":"Heart & Blood Pressure","C03":"Heart & Blood Pressure",
    "C04":"Heart & Blood Pressure","C05":"Heart & Blood Pressure","C07":"Heart & Blood Pressure",
    "C08":"Heart & Blood Pressure","C09":"Heart & Blood Pressure","C10":"Cholesterol",
    "D01":"Antifungals","D02":"Skin & Wounds","D03":"Skin & Wounds","D04":"Skin & Wounds",
    "D05":"Skin & Wounds","D06":"Antibiotics","D07":"Corticosteroids","D08":"Skin & Wounds",
    "D09":"Skin & Wounds","D10":"Skin & Wounds","D11":"Skin & Wounds",
    "G01":"Women's Health","G02":"Women's Health","G03":"Women's Health","G04":"Urology",
    "H01":"Thyroid","H02":"Corticosteroids","H03":"Thyroid","H04":"Diabetes",
    "H05":"Vitamins & Supplements",
    "J01":"Antibiotics","J02":"Antifungals","J04":"Antibiotics","J05":"Antivirals",
    "J06":"Antivirals","J07":"Antivirals",
    "L01":"Oncology","L02":"Oncology","L03":"Oncology","L04":"Corticosteroids",
    "M01":"Pain & Fever","M02":"Joints & Muscles","M03":"Joints & Muscles",
    "M04":"Joints & Muscles","M05":"Joints & Muscles","M09":"Joints & Muscles",
    "N01":"Pain & Fever","N02":"Pain & Fever","N03":"Neurology","N04":"Neurology",
    "N05":"Sleep & Sedation","N06":"Antidepressants","N07":"Nervous System",
    "P01":"Antiparasitics","P02":"Antiparasitics","P03":"Antiparasitics",
    "R01":"Cough & Cold","R02":"Cough & Cold","R03":"Lungs & Asthma",
    "R04":"Cough & Cold","R05":"Cough & Cold","R06":"Allergy","R07":"Lungs & Asthma",
    "S01":"Eye & Ear","S02":"Eye & Ear","S03":"Eye & Ear",
    "V03":"First Aid","V06":"Vitamins & Supplements","V07":"First Aid","V08":"First Aid",
}

BLACKLIST = re.compile(
    r"\b(vaccine|vaccin|immunoglobulin|albumin|dialys|diagnostic|device|veterinär|veterinary|radiopharm)\b", re.I
)
WITHDRAWN = re.compile(
    r"withdrawn|refused|suspended|expired|revoked|tilbagekaldt|retirado|zurückgezogen|ritirato|wycofany|retiré", re.I
)

def atc_cat(atc):
    return ATC_MAP.get((atc or "").strip()[:3].upper())

def curl(url, dest, max_time=120):
    cmd = ["curl","-L","--max-time",str(max_time),"--connect-timeout","20",
           "--silent","--fail","--user-agent","Mozilla/5.0 apoHouze-updater/5.0",
           "-o",dest,url]
    for i in range(3):
        try:
            subprocess.run(cmd,timeout=max_time+15,check=True)
            size = os.path.getsize(dest)
            print(f"  ✅ {size//1024} KB")
            return size
        except Exception as e:
            print(f"  ⚠️  Poging {i+1}/3: {e}")
            if i<2: time.sleep(4)
    return 0

def is_html(path):
    """Controleer of bestand eigenlijk HTML is (foutpagina)."""
    try:
        with open(path,"rb") as f:
            head = f.read(512).lower()
        return b"<!doctype" in head or b"<html" in head
    except:
        return False

def http_json(url, timeout=20):
    req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0 apoHouze/1.0","Accept":"application/json"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read())

# ─── EMA JSON (gedeelde basis voor alle EU-landen) ───────────────
def fetch_ema(seen):
    print("  📥 EMA JSON...")
    dest = os.path.join(TMP_DIR,"shared_ema.json")
    if os.path.exists(dest) and (time.time()-os.path.getmtime(dest))<14400:
        print("  ♻️  Gecachede EMA-data gebruiken")
    else:
        if curl("https://www.ema.europa.eu/en/documents/report/medicines-output-medicines_json-report_en.json",dest) < 10000:
            return []
    try:
        with open(dest,encoding="utf-8-sig") as f: raw = json.load(f)
    except:
        return []
    items = raw if isinstance(raw,list) else next((v for v in raw.values() if isinstance(v,list)),[])
    keys = list(items[0].keys()) if items else []
    def find(*pats):
        for k in keys:
            kl = k.lower().replace(" ","_")
            if any(re.search(p,kl) for p in pats): return k
        return None
    nk = find(r"medicine.*name",r"name_of_medicine",r"^name$")
    ik = find(r"active_substance",r"\binn\b")
    ak = find(r"^atc")
    sk = find(r"authoris",r"status")
    results,sk_cnt = [],0
    for item in items:
        name = str(item.get(nk) or "").strip()
        if not name: continue
        status = str(item.get(sk) or "") if sk else ""
        if status and WITHDRAWN.search(status): sk_cnt+=1; continue
        if BLACKLIST.search(name): sk_cnt+=1; continue
        atc = str(item.get(ak) or "").strip() if ak else ""
        inn = str(item.get(ik) or "").strip() if ik else ""
        cat = atc_cat(atc)
        if not cat: sk_cnt+=1; continue
        key = name.lower()
        if key in seen: continue
        seen.add(key)
        results.append({"Name":name,"INN":inn,"ATC":atc,"PharmaceuticalForm":"","RxStatus":"Rx","Country":COUNTRY,"Category":cat})
    print(f"  ✅ {len(results)} EMA | {sk_cnt} overgeslagen")
    return results

# ─── Generieke CSV parser ─────────────────────────────────────────
def find_col(sample, patterns):
    """Zoek kolom op basis van patronen. Slaat None-sleutels over."""
    for k in sample.keys():
        if k is None: continue          # ← FIX: None-sleutels skippen
        kl = k.lower().strip()
        if any(re.search(p, kl) for p in patterns):
            return k
    return None

def parse_csv(path, country, seen, name_pats, inn_pats, atc_pats,
              form_pats=None, status_pats=None, rx_pats=None, sep=None):
    if is_html(path):
        print("  ⚠️  Bestand is HTML (foutpagina) - overgeslagen")
        return []
    for enc in ("utf-8-sig","latin-1","utf-8","cp1250","cp1252"):
        try:
            with open(path,encoding=enc,errors="strict") as f:
                sample = f.read(8192); f.seek(0)
                if sep is None:
                    counts = {s:sample.count(s) for s in (";","\t",",")}
                    detected_sep = max(counts,key=counts.get)
                else:
                    detected_sep = sep
                reader = csv.DictReader(f,delimiter=detected_sep)
                rows = list(reader)
            break
        except UnicodeDecodeError:
            continue
    else:
        return []
    if not rows: return []
    print(f"  📊 {len(rows)} rijen | sep='{detected_sep}' | enc={enc}")
    s = rows[0]
    nk  = find_col(s,name_pats)
    ik  = find_col(s,inn_pats)
    ak  = find_col(s,atc_pats)
    fk  = find_col(s,form_pats) if form_pats else None
    stk = find_col(s,status_pats) if status_pats else None
    rxk = find_col(s,rx_pats) if rx_pats else None
    if not nk and not ik:
        print(f"  ⚠️  Geen naam/INN kolom. Kolommen: {[k for k in s.keys() if k][:8]}")
        return []
    results,sk_bl,sk_cat,sk_st,sk_dup = [],0,0,0,0
    for row in rows:
        name   = str(row.get(nk) or "").strip() if nk else ""
        inn    = str(row.get(ik) or "").strip()  if ik else ""
        atc    = str(row.get(ak) or "").strip()  if ak else ""
        form   = str(row.get(fk) or "").strip()  if fk else ""
        status = str(row.get(stk) or "").strip() if stk else ""
        rxraw  = str(row.get(rxk) or "").strip() if rxk else ""
        display = name or inn
        if not display: continue
        if BLACKLIST.search(display): sk_bl+=1; continue
        if status and WITHDRAWN.search(status): sk_st+=1; continue
        cat = atc_cat(atc)
        if not cat: sk_cat+=1; continue
        rx = bool(re.search(r"recept|prescri|liste[_ ]?i|verschreib|resepti|verkrijg",rxraw,re.I))
        key = display.lower()
        if key in seen: sk_dup+=1; continue
        seen.add(key)
        results.append({"Name":display,"INN":inn,"ATC":atc,"PharmaceuticalForm":form,
                        "RxStatus":"Rx" if rx else "OTC","Country":country,"Category":cat or ""})
    print(f"  ✅ {len(results)} | bl:{sk_bl} cat:{sk_cat} st:{sk_st} dup:{sk_dup}")
    return results

# ─── Landspecifieke fetchers ──────────────────────────────────────

def fetch_es(seen):
    """Spanje — AEMPS CIMA REST API
    Max 100 pagina's (10000 records) om timeout te vermijden.
    Elke pagina heeft een eigen timeout van 10s.
    """
    print("  📥 AEMPS CIMA (Spanje)...")
    results = []; page = 1; max_pages = 100
    consecutive_errors = 0
    while page <= max_pages:
        url = f"https://cima.aemps.es/cima/rest/medicamentos?estado=1&pageSize=100&numPagina={page}"
        try:
            data = http_json(url, timeout=10)
            consecutive_errors = 0
        except Exception as e:
            consecutive_errors += 1
            print(f"  ⚠️  CIMA p{page}: {e}")
            if consecutive_errors >= 3:
                print(f"  ⚠️  3 opeenvolgende fouten - stoppen")
                break
            time.sleep(2)
            continue
        items = data.get("resultados",[])
        if not items:
            if page == 1:
                print(f"  ⚠️  CIMA: pagina 1 leeg (API mogelijk rate-limited)")
            break
        for m in items:
            name = (m.get("nombre") or "").strip()
            inn  = ((m.get("principiosActivos") or [{}])[0].get("nombre","") if m.get("principiosActivos") else "")
            atc  = ((m.get("atcs") or [{}])[0].get("codigo","") if m.get("atcs") else "")
            form = (m.get("formaFarmaceutica") or {}).get("nombre","")
            if not name or BLACKLIST.search(name): continue
            cat = atc_cat(atc)
            if not cat: continue
            key = name.lower()
            if key in seen: continue
            seen.add(key)
            results.append({"Name":name,"INN":inn,"ATC":atc,"PharmaceuticalForm":form,
                            "RxStatus":"Rx","Country":"ES","Category":cat})
        total = data.get("totalFilas",0)
        if total and page*100 >= total: break
        page += 1
        time.sleep(0.3)
    print(f"  ✅ {len(results)} CIMA-producten")
    return results

def fetch_de(seen):
    """Duitsland — BfArM Dringlichkeitsliste + pädiatrische lijst"""
    print("  📥 BfArM (Duitsland)...")
    results = []
    # Dringlichkeitsliste (lijst van kritische geneesmiddelen met alle vergunninghouders)
    # Dit bestand bevat: Bezeichnung, ATC, PZN, Wirkstoff, Darreichungsform
    urls = [
        "https://www.bfarm.de/SharedDocs/Downloads/DE/Arzneimittel/Zulassung/lieferengpaesse/dringlichkeitsliste_arzneimittelliste.csv",
        "https://www.bfarm.de/SharedDocs/Downloads/DE/Arzneimittel/Zulassung/amInformationen/Lieferengpaesse/dringlichkeitsliste_arzneimittelliste.csv",
    ]
    dest = os.path.join(TMP_DIR,"de_dringlichkeit.csv")
    for url in urls:
        print(f"  📥 {url}")
        if curl(url,dest) > 1000 and not is_html(dest):
            r = parse_csv(dest,"DE",seen,
                name_pats=[r"bezeichnung|handelsname|name|produkt"],
                inn_pats=[r"wirkstoff|inn|substanz"],
                atc_pats=[r"^atc"],
                form_pats=[r"darreich|form"],
                status_pats=[r"status|zulass"],
                rx_pats=[r"verschreib|rezept"])
            results.extend(r)
            break
        print(f"  ⚠️  Niet bereikbaar")
    # Pädiatrische lijst (essentiële kindergeneesmiddelen met ATC)
    paed_urls = [
        "https://www.bfarm.de/SharedDocs/Downloads/DE/Arzneimittel/Zulassung/lieferengpaesse/liste_paediatrische_arzneimittel.csv",
        "https://www.bfarm.de/SharedDocs/Downloads/DE/Arzneimittel/Zulassung/amInformationen/Lieferengpaesse/liste_paediatrische_arzneimittel.csv",
    ]
    dest2 = os.path.join(TMP_DIR,"de_paed.csv")
    for url in paed_urls:
        print(f"  📥 {url}")
        if curl(url,dest2) > 1000 and not is_html(dest2):
            r = parse_csv(dest2,"DE",seen,
                name_pats=[r"bezeichnung|handelsname|name|produkt"],
                inn_pats=[r"wirkstoff|inn|substanz"],
                atc_pats=[r"^atc"],
                form_pats=[r"darreich|form"])
            results.extend(r)
            break
    print(f"  ✅ {len(results)} BfArM-producten totaal")
    return results

def fetch_se(seen):
    """Zweden — Läkemedelsverket tekortenlijst (shortage list)"""
    print("  📥 Läkemedelsverket (Zweden)...")
    dest = os.path.join(TMP_DIR,"se_lv.csv")
    urls = [
        # Läkemedelsverket - shortage lijst
        "https://www.lakemedelsverket.se/globalassets/produkt-och-tillstand/lakemedelsregistret/lakemedel-med-avbrutet-godkannande.csv",
        "https://www.lakemedelsverket.se/globalassets/produkt-och-tillstand/lakemedelsregistret/lakemedelsstatistik.csv",
        # TLV (Tandvårds- och läkemedelsförmånsverket) - vergoedde geneesmiddelen CSV
        "https://www.tlv.se/download/18.5b4ad01917a3a6a0e3c6c5/1651745067789/periodens_vara.csv",
        "https://www.tlv.se/download/18.47be06b917a3a6c59d38bff/1616510649766/beslutslista.csv",
        # dataportal.se CKAN voor LV datasets
        "https://dataportal.se/api/3/action/package_search?q=lakemedel+lakemedelsverket&rows=5",
    ]
    for url in urls:
        print(f"  📥 {url}")
        if url.endswith(".json") or "ckan" in url or "api/3" in url:
            try:
                data = http_json(url, timeout=15)
                pkgs = data.get("result",{}).get("results",[])
                for pkg in pkgs:
                    for r in pkg.get("resources",[]):
                        if ".csv" in r.get("url","").lower():
                            size = curl(r["url"],dest)
                            if size > 5000 and not is_html(dest):
                                return parse_csv(dest,"SE",seen,
                                    name_pats=[r"produktnamn|l.kemedel.*namn|name|handelsnamn"],
                                    inn_pats=[r"substans|aktiv.*ingred|inn"],
                                    atc_pats=[r"^atc"],
                                    form_pats=[r"l.kemedels.*form|form"],
                                    status_pats=[r"status|godkänn"],
                                    rx_pats=[r"recept|förskriv"])
            except Exception as e:
                print(f"  ⚠️  {e}")
            continue
        if curl(url,dest) > 5000 and not is_html(dest):
            return parse_csv(dest,"SE",seen,
                name_pats=[r"produktnamn|l.kemedel.*namn|name"],
                inn_pats=[r"substans|aktiv.*ingred|inn"],
                atc_pats=[r"^atc"],
                form_pats=[r"form"],
                status_pats=[r"status"],
                rx_pats=[r"recept"])
        print(f"  ⚠️  Niet bereikbaar")
    return []

def fetch_no(seen):
    """Noorwegen — NoMA Statens legemiddelverk (shortage list CSV)"""
    print("  📥 NoMA (Noorwegen)...")
    dest = os.path.join(TMP_DIR,"no_noma.csv")
    urls = [
        # NoMA shortage lijst (leveringsproblemer)
        "https://www.legemiddelverket.no/globalassets/tilgjengelighet-og-pris/leveringsproblemer/leveringsproblemer.xlsx",
        "https://www.legemiddelverket.no/globalassets/tilgjengelighet-og-pris/leveringsproblemer/leveringsproblemer.csv",
        # Folkehelseinstituttet statistikk - legemiddelstatistikk CSV
        "https://www.fhi.no/contentassets/8c65b49d03924aee827b31ebbf71a0a3/legemiddelstatistikk.csv",
        # data.norge.no CKAN
        "https://data.norge.no/api/3/action/package_search?q=legemidler&rows=5",
    ]
    for url in urls:
        print(f"  📥 {url}")
        if curl(url,dest) > 5000 and not is_html(dest):
            if url.endswith(".xlsx"):
                return parse_xlsx(dest,"NO",seen)
            return parse_csv(dest,"NO",seen,
                name_pats=[r"varenavn|produktnavn|name|handelsnavn"],
                inn_pats=[r"virkestoff|substans|inn"],
                atc_pats=[r"^atc"],
                form_pats=[r"legemiddelform|form"],
                status_pats=[r"status"],
                rx_pats=[r"resept"])
        print(f"  ⚠️  Niet bereikbaar")
    return []

def fetch_dk(seen):
    """Denemarken — Lægemiddelstyrelsen (shortage list)"""
    print("  📥 Lægemiddelstyrelsen (Denemarken)...")
    dest = os.path.join(TMP_DIR,"dk_dkma.csv")
    urls = [
        # DKMA shortage lijst (leveringssvigt) - directe XLSX/CSV links
        "https://laegemiddelstyrelsen.dk/globalassets/sundhedspersoner/medicinsk-udstyr/tilgaengelighed/leveringssvigt/leveringssvigt.xlsx",
        "https://laegemiddelstyrelsen.dk/globalassets/sundhedspersoner/tilgaengelighed-af-lagemidler/leveringssvigt.xlsx",
        # medicinpriser.dk download - werkt al maar crasht door None-kolom
        "https://www.medicinpriser.dk/default.aspx?action=downloadfile&file=medicineprices.csv",
        # Sundhedsdatastyrelsen
        "https://sundhedsdatastyrelsen.dk/-/media/sds/filer/rammer-og-retningslinjer/takster-og-priser/medicinpriser/medicinpriser.zip",
    ]
    for url in urls:
        print(f"  📥 {url}")
        size = curl(url,dest)
        if size < 1000:
            print(f"  ⚠️  Niet bereikbaar"); continue
        if is_html(dest):
            print(f"  ⚠️  HTML-foutpagina"); continue
        if url.endswith(".xlsx"):
            r = parse_xlsx(dest,"DK",seen); 
            if r: return r
        elif url.endswith(".zip"):
            r = parse_zip(dest,"DK",seen)
            if r: return r
        else:
            r = parse_csv(dest,"DK",seen,
                name_pats=[r"produktnavn|varenavn|l.gemiddel.*navn|name"],
                inn_pats=[r"virkestof|substans|inn"],
                atc_pats=[r"^atc"],
                form_pats=[r"form"],
                status_pats=[r"status"],
                rx_pats=[r"recept"])
            if r: return r
    return []

def fetch_fi(seen):
    """Finland — Fimea (shortage list)"""
    print("  📥 Fimea (Finland)...")
    dest = os.path.join(TMP_DIR,"fi_fimea.csv")
    urls = [
        # Fimea publiceert lääkevalmisteiden saatavuushäiriö (shortage) lijst
        "https://www.fimea.fi/documents/160140/0/saatavuushairio.csv",
        "https://www.fimea.fi/documents/160140/753095/saatavuushairio.csv",
        # avoindata.fi CKAN
    ]
    # Probeer avoindata.fi CKAN voor Fimea + THL datasets
    try:
        data = http_json("https://www.avoindata.fi/api/3/action/package_search?q=laake+fimea&rows=10",timeout=15)
        pkgs = data.get("result",{}).get("results",[])
        for pkg in pkgs:
            for r in pkg.get("resources",[]):
                if ".csv" in r.get("url","").lower() or r.get("format","").upper()=="CSV":
                    urls.insert(0, r["url"])
    except Exception as e:
        print(f"  ⚠️  avoindata.fi: {e}")
    for url in urls:
        print(f"  📥 {url}")
        if curl(url,dest) > 1000 and not is_html(dest):
            return parse_csv(dest,"FI",seen,
                name_pats=[r"kauppanimi|nimi|name|valmisteen.*nimi"],
                inn_pats=[r"vaikuttava.*aine|substans|inn|api"],
                atc_pats=[r"^atc"],
                form_pats=[r"l..kemuoto|form"],
                status_pats=[r"status|myynti"])
        print(f"  ⚠️  Niet bereikbaar")
    return []

def fetch_at(seen):
    """Oostenrijk — AGES/BASG (shortage list via data.gv.at)"""
    print("  📥 BASG (Oostenrijk)...")
    dest = os.path.join(TMP_DIR,"at_ages.csv")
    # Probeer data.gv.at CKAN voor AGES datasets
    urls = []
    try:
        data = http_json("https://www.data.gv.at/katalog/api/3/action/package_search?q=arzneimittel+ages&rows=10",timeout=15)
        pkgs = data.get("result",{}).get("results",[])
        for pkg in pkgs:
            for r in pkg.get("resources",[]):
                url = r.get("url","")
                if any(ext in url.lower() for ext in [".csv",".xlsx",".zip"]):
                    urls.append(url)
    except Exception as e:
        print(f"  ⚠️  data.gv.at: {e}")
    # Directe AGES URLs
    urls += [
        "https://www.ages.at/download/0/0/lieferengpaesse.csv",
        "https://www.ages.at/fileadmin/AGES2015/Themen/Arzneimittel/lieferengpaesse.csv",
        "https://www.basg.gv.at/fileadmin/user_upload/060_Arzneimittel/lieferengpaesse.csv",
        "https://data.gv.at/katalog/api/3/action/package_search?q=arzneimittel+ages&rows=5",
    ]
    for url in urls[:8]:  # max 8 pogingen
        print(f"  📥 {url}")
        if curl(url,dest) > 5000 and not is_html(dest):
            if url.endswith(".xlsx"):
                return parse_xlsx(dest,"AT",seen)
            return parse_csv(dest,"AT",seen,
                name_pats=[r"bezeichnung|handelsname|name|arzneimittel"],
                inn_pats=[r"wirkstoff|substanz|inn"],
                atc_pats=[r"^atc"],
                form_pats=[r"darreich|form"],
                status_pats=[r"status|zulass"],
                rx_pats=[r"abgabe|verschreib"])
        print(f"  ⚠️  Niet bereikbaar")
    return []

def fetch_ch(seen):
    """Zwitserland — Swissmedic shortage list"""
    print("  📥 Swissmedic (Zwitserland)...")
    dest = os.path.join(TMP_DIR,"ch_sm.xlsx")
    urls = [
        "https://www.swissmedic.ch/dam/swissmedic/de/dokumente/humanarzneimittel/versorgungsengpaesse/liste_versorgungsengpaesse.xlsx.download.xlsx/Liste_Versorgungsengpaesse.xlsx",
        "https://www.swissmedic.ch/dam/swissmedic/de/dokumente/humanarzneimittel/versorgungsengpaesse/versorgungsengpaesse.xlsx.download.xlsx/Versorgungsengpaesse.xlsx",
        "https://www.swissmedic.ch/swissmedic/de/home/humanarzneimittel/versorgungsengpaesse.html.downloadliste.xlsx",
    ]
    for url in urls:
        print(f"  📥 {url}")
        if curl(url,dest) > 5000 and not is_html(dest):
            return parse_xlsx(dest,"CH",seen)
        print(f"  ⚠️  Niet bereikbaar")
    return []

def fetch_ie(seen):
    """Ierland — HPRA shortage list / data.gov.ie"""
    print("  📥 HPRA (Ierland)...")
    dest = os.path.join(TMP_DIR,"ie_hpra.csv")
    # Probeer data.gov.ie CKAN
    urls = []
    try:
        data = http_json("https://data.gov.ie/api/3/action/package_search?q=medicines+hpra&rows=5",timeout=15)
        for pkg in data.get("result",{}).get("results",[]):
            for r in pkg.get("resources",[]):
                if ".csv" in r.get("url","").lower() or r.get("format","").upper()=="CSV":
                    urls.append(r["url"])
    except Exception as e:
        print(f"  ⚠️  data.gov.ie: {e}")
    urls += [
        "https://www.hpra.ie/docs/default-source/default-document-library/medicine-shortage.csv",
        "https://www.hpra.ie/docs/default-source/shortages/current-medicine-shortages.csv",
        "https://www.hpra.ie/docs/default-source/default-document-library/shortage-register.csv",
    ]
    for url in urls[:6]:
        print(f"  📥 {url}")
        if curl(url,dest) > 5000 and not is_html(dest):
            return parse_csv(dest,"IE",seen,
                name_pats=[r"product|name|medicine"],
                inn_pats=[r"active|inn|substance"],
                atc_pats=[r"^atc"],
                form_pats=[r"form"],
                status_pats=[r"status"],
                rx_pats=[r"prescription"])
        print(f"  ⚠️  Niet bereikbaar")
    return []

def fetch_pt(seen):
    """Portugal — INFARMED (shortage list)"""
    print("  📥 INFARMED (Portugal)...")
    dest = os.path.join(TMP_DIR,"pt_infarmed.csv")
    urls = [
        # INFARMED shortage list (indisponibilidades) - directe document IDs
        "https://www.infarmed.pt/documents/15786/17838/indisponibilidades.csv",
        "https://www.infarmed.pt/documents/15786/1929865/indisponibilidades.csv",
        "https://www.infarmed.pt/web/infarmed/indisponibilidades-de-medicamentos-temporarias",
        # dados.gov.pt CKAN voor INFARMED datasets
        "https://dados.gov.pt/api/3/action/package_search?q=infarmed+medicamentos&rows=5",
    ]
    for url in urls:
        print(f"  📥 {url}")
        if curl(url,dest) > 5000 and not is_html(dest):
            return parse_csv(dest,"PT",seen,
                name_pats=[r"denominação|nome.*medic|nome.*comerc|name"],
                inn_pats=[r"denomin.*comum|substância|dci|inn"],
                atc_pats=[r"^atc"],
                form_pats=[r"forma.*farm|form"],
                status_pats=[r"estado|situação"])
        print(f"  ⚠️  Niet bereikbaar")
    return []

def fetch_pl(seen):
    """Polen — URPL shortage list / GIF"""
    print("  📥 URPL (Polen)...")
    dest = os.path.join(TMP_DIR,"pl_urpl.csv")
    urls = [
        # GIF shortage list (Braki Produktów Leczniczych)
        "https://www.gif.gov.pl/download/3/21813/wykazproduktowalarmowych.csv",
        "https://www.gif.gov.pl/pl/rejestry/produkty-lecznicze/wykazproduktowalarmowych.csv",
        # URPL bulk export - andere endpoints
        "https://rejestry.ezdrowie.gov.pl/api/rpl/medicinal-products/shortage-list.csv",
        "https://rejestry.ezdrowie.gov.pl/rpl/api/public/v1/products/export?format=csv",
        # dane.gov.pl CKAN voor URPL datasets
        "https://api.dane.gov.pl/api/3/action/package_search?q=produkty+lecznicze+URPL&rows=5",
    ]
    for url in urls:
        print(f"  📥 {url}")
        if curl(url,dest) > 5000 and not is_html(dest):
            return parse_csv(dest,"PL",seen,
                name_pats=[r"nazwa.*produktu|nazwa.*handl|name"],
                inn_pats=[r"substancja.*czynna|inn|substancja"],
                atc_pats=[r"^atc|kod.*atc"],
                form_pats=[r"postać|forma"])
        print(f"  ⚠️  Niet bereikbaar")
    return []

# ─── XLSX parser ─────────────────────────────────────────────────
def parse_xlsx(path, country, seen,
               name_pats=None, inn_pats=None, atc_pats=None,
               form_pats=None, status_pats=None, rx_pats=None):
    if name_pats is None: name_pats=[r"bezeichnung|handelsname|name|produkt|varenavn|produktnamn|nom"]
    if inn_pats  is None: inn_pats =[r"wirkstoff|substanz|inn|virkestof|substans|dci"]
    if atc_pats  is None: atc_pats =[r"^atc"]
    if form_pats is None: form_pats=[r"darreich|form|l.kemedels.*form"]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path,read_only=True,data_only=True)
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows_raw: return []
        headers = [str(h).strip() if h is not None else f"col{i}" for i,h in enumerate(rows_raw[0])]
        rows = [{headers[i]:(str(rows_raw[r][i]).strip() if i<len(rows_raw[r]) and rows_raw[r][i] is not None else "")
                 for i in range(len(headers))} for r in range(1,len(rows_raw))]
        print(f"  📊 {len(rows)} rijen (xlsx)")
        return parse_csv.__wrapped__(rows,country,seen,name_pats,inn_pats,atc_pats,form_pats,status_pats,rx_pats) if hasattr(parse_csv,'__wrapped__') else _process_rows(rows,country,seen,name_pats,inn_pats,atc_pats,form_pats,status_pats,rx_pats)
    except Exception as e:
        print(f"  ⚠️  XLSX fout: {e}"); return []

def _process_rows(rows, country, seen, name_pats, inn_pats, atc_pats,
                  form_pats=None, status_pats=None, rx_pats=None):
    if not rows: return []
    s = rows[0]
    nk  = find_col(s,name_pats)
    ik  = find_col(s,inn_pats)
    ak  = find_col(s,atc_pats)
    fk  = find_col(s,form_pats) if form_pats else None
    stk = find_col(s,status_pats) if status_pats else None
    rxk = find_col(s,rx_pats) if rx_pats else None
    if not nk and not ik:
        print(f"  ⚠️  Geen kolom. Kolommen: {[k for k in s.keys() if k][:8]}"); return []
    results,sk_bl,sk_cat,sk_dup = [],0,0,0
    for row in rows:
        name = str(row.get(nk) or "").strip() if nk else ""
        inn  = str(row.get(ik) or "").strip()  if ik else ""
        atc  = str(row.get(ak) or "").strip()  if ak else ""
        form = str(row.get(fk) or "").strip()  if fk else ""
        status = str(row.get(stk) or "").strip() if stk else ""
        display = name or inn
        if not display: continue
        if BLACKLIST.search(display): sk_bl+=1; continue
        if status and WITHDRAWN.search(status): continue
        cat = atc_cat(atc)
        if not cat: sk_cat+=1; continue
        key = display.lower()
        if key in seen: sk_dup+=1; continue
        seen.add(key)
        results.append({"Name":display,"INN":inn,"ATC":atc,"PharmaceuticalForm":form,
                        "RxStatus":"Rx","Country":country,"Category":cat or ""})
    print(f"  ✅ {len(results)} xlsx-entries | cat:{sk_cat} dup:{sk_dup}")
    return results

# Monkey-patch xlsx om _process_rows te gebruiken
def parse_xlsx(path, country, seen, **kw):
    name_pats = kw.get("name_pats",[r"bezeichnung|handelsname|name|produkt|varenavn|produktnamn|nom|nimi"])
    inn_pats  = kw.get("inn_pats", [r"wirkstoff|substanz|inn|virkestof|substans|dci|vaikuttava"])
    atc_pats  = kw.get("atc_pats", [r"^atc"])
    form_pats = kw.get("form_pats",[r"darreich|form"])
    status_pats=kw.get("status_pats",[r"status|zulass|godkänn"])
    rx_pats   = kw.get("rx_pats",  [r"recept|verschreib|resepti"])
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path,read_only=True,data_only=True)
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows_raw: return []
        headers = [str(h).strip() if h is not None else f"col{i}" for i,h in enumerate(rows_raw[0])]
        rows = [{headers[i]:(str(rows_raw[r][i]).strip() if i<len(rows_raw[r]) and rows_raw[r][i] is not None else "")
                 for i in range(len(headers))} for r in range(1,len(rows_raw))]
        print(f"  📊 {len(rows)} rijen (xlsx)")
        return _process_rows(rows,country,seen,name_pats,inn_pats,atc_pats,form_pats,status_pats,rx_pats)
    except Exception as e:
        print(f"  ⚠️  XLSX fout: {e}"); return []

def parse_zip(path, country, seen):
    """Pak een ZIP uit en verwerk de eerste CSV/XLSX die erin zit."""
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            for name in names:
                if name.lower().endswith(".csv") or name.lower().endswith(".txt"):
                    inner = os.path.join(TMP_DIR,f"inner_{country.lower()}.csv")
                    with z.open(name) as zf, open(inner,"wb") as out:
                        out.write(zf.read())
                    if not is_html(inner):
                        return parse_csv(inner,country,seen,
                            name_pats=[r"produktnavn|varenavn|name|produkt"],
                            inn_pats=[r"virkestof|substans|inn"],
                            atc_pats=[r"^atc"],
                            form_pats=[r"form"])
    except Exception as e:
        print(f"  ⚠️  ZIP fout: {e}")
    return []

# ─── Dispatcher ───────────────────────────────────────────────────
FETCHERS = {
    "ES": fetch_es, "DE": fetch_de,
    "AT": fetch_at, "CH": fetch_ch,
    "DK": fetch_dk, "FI": fetch_fi,
    "IE": fetch_ie, "NO": fetch_no,
    "PL": fetch_pl, "PT": fetch_pt,
    "SE": fetch_se,
}

FLAG = {"AT":"🇦🇹","CH":"🇨🇭","DK":"🇩🇰","ES":"🇪🇸","FI":"🇫🇮",
        "IE":"🇮🇪","NO":"🇳🇴","PL":"🇵🇱","PT":"🇵🇹","SE":"🇸🇪","DE":"🇩🇪"}

def main():
    print(f"{FLAG.get(COUNTRY,'🌍')} apoHouze — {COUNTRY} Medicijnen Fetcher v2")
    print("="*52)
    seen = set()
    all_results = []
    print("\n[1/2] EMA JSON...")
    all_results.extend(fetch_ema(seen))
    if COUNTRY in FETCHERS:
        print(f"\n[2/2] Nationale bron ({COUNTRY})...")
        all_results.extend(FETCHERS[COUNTRY](seen))
    print(f"\n  🎯 Totaal: {len(all_results)} unieke medicijnen")
    if not all_results:
        print("❌ Geen resultaten"); sys.exit(1)
    fields = ["Name","INN","ATC","PharmaceuticalForm","RxStatus","Country","Category"]
    with open(OUTPUT,"w",newline="",encoding="utf-8") as f:
        w = csv.DictWriter(f,fieldnames=fields,extrasaction="ignore")
        w.writeheader(); w.writerows(all_results)
    print(f"✅ {len(all_results)} opgeslagen → {OUTPUT}")

if __name__ == "__main__":
    main()
